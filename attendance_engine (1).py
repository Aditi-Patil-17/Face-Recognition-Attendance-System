"""
=============================================================
  ATTENDANCE ENGINE — No-Stop Attendance System
  Features: Entry + Exit tracking, Excel logging, Debug mode
=============================================================
USAGE:
    python attendance_engine.py

INSTALL:
    pip install deepface tf-keras opencv-python openpyxl numpy
=============================================================
"""

import os
os.environ['TF_ENABLE_ONEDNN_OPTS'] = '0'
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '3'

import cv2
import time
import numpy as np
import openpyxl
from datetime import datetime
from deepface import DeepFace

# ── CONFIG ────────────────────────────────────────────────
KNOWN_FACES_DIR    = "known_faces"
EXCEL_FILE         = "attendance.xlsx"
CAMERA_INDEX       = 0
MODEL_NAME         = "Facenet"          # faster than ArcFace
DETECTOR_BACKEND   = "opencv"
DISTANCE_THRESHOLD = 0.50              # lower = stricter
FRAME_SKIP         = 10               # process every 10th frame

# Exit detection config
# If a person was seen, then NOT seen for this many seconds → log as exited
EXIT_TIMEOUT_SECONDS = 15

# ── EXCEL SETUP ───────────────────────────────────────────
def init_excel(path):
    if not os.path.exists(path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance"
        ws.append(["Name", "Date", "Entry Time", "Exit Time", "Duration", "Status"])
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["F"].width = 10
        wb.save(path)
        print(f"  [EXCEL] Created new file: {path}")

def find_todays_row(ws, name):
    """Returns row number if person already has an entry today, else None."""
    today = datetime.now().strftime("%Y-%m-%d")
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] == name and str(row[1]) == today:
            return i
    return None

def log_entry(path, name):
    """Log entry time. Returns True if newly logged."""
    wb  = openpyxl.load_workbook(path)
    ws  = wb.active
    row = find_todays_row(ws, name)
    if row is not None:
        wb.close()
        return False  # already logged entry today

    now    = datetime.now()
    status = "On-time" if (now.hour < 9 or (now.hour == 9 and now.minute <= 15)) else "Late"
    ws.append([name, now.strftime("%Y-%m-%d"),
               now.strftime("%H:%M:%S"), "", "", status])
    wb.save(path)
    print(f"  [ENTRY] {name} entered at {now.strftime('%H:%M:%S')} — {status}")
    return True

def log_exit(path, name):
    """Fill in exit time and duration for today's entry row."""
    wb   = openpyxl.load_workbook(path)
    ws   = wb.active
    row  = find_todays_row(ws, name)
    if row is None:
        wb.close()
        return

    now        = datetime.now()
    entry_cell = ws.cell(row=row, column=3).value   # Entry Time
    if entry_cell:
        try:
            today      = datetime.now().strftime("%Y-%m-%d")
            entry_dt   = datetime.strptime(f"{today} {entry_cell}", "%Y-%m-%d %H:%M:%S")
            duration   = now - entry_dt
            hours, rem = divmod(int(duration.total_seconds()), 3600)
            minutes    = rem // 60
            dur_str    = f"{hours}h {minutes}m"
        except:
            dur_str = ""
    else:
        dur_str = ""

    ws.cell(row=row, column=4).value = now.strftime("%H:%M:%S")
    ws.cell(row=row, column=5).value = dur_str
    wb.save(path)
    print(f"  [EXIT]  {name} exited at {now.strftime('%H:%M:%S')} — Duration: {dur_str}")

# ── LOAD KNOWN FACES ──────────────────────────────────────
def load_known_faces(directory):
    known = {}

    if not os.path.exists(directory):
        print(f"  [ERROR] Folder not found: {os.path.abspath(directory)}")
        return known

    entries = os.listdir(directory)
    print(f"  [DEBUG] Contents of known_faces/: {entries}")

    for entry in entries:
        entry_path = os.path.join(directory, entry)

        # Support BOTH structures:
        # Structure A: known_faces/Aditi_Sharma/front_01.jpg  (subfolder per person)
        # Structure B: known_faces/Aditi_Sharma.jpg           (flat image per person)

        if os.path.isdir(entry_path):
            # Structure A — subfolder
            name   = entry
            images = [f for f in os.listdir(entry_path)
                      if f.lower().endswith((".jpg",".jpeg",".png"))]
            print(f"  [DEBUG] '{name}' folder has {len(images)} images: {images[:5]}")
            embeddings = []
            for img_file in images:
                img_path = os.path.join(entry_path, img_file)
                emb = get_embedding(img_path)
                if emb is not None:
                    embeddings.append(emb)
            if embeddings:
                known[name] = embeddings
                print(f"  [OK] Loaded {len(embeddings)} embeddings for: {name}")
            else:
                print(f"  [WARN] No valid embeddings found for: {name}")

        elif entry.lower().endswith((".jpg",".jpeg",".png")):
            # Structure B — flat image
            name = os.path.splitext(entry)[0]
            emb  = get_embedding(entry_path)
            if emb is not None:
                known[name] = [emb]
                print(f"  [OK] Loaded 1 embedding for: {name}")

    return known

def get_embedding(img_path):
    try:
        result = DeepFace.represent(
            img_path          = img_path,
            model_name        = MODEL_NAME,
            detector_backend  = DETECTOR_BACKEND,
            enforce_detection = False
        )
        if result and len(result) > 0:
            return np.array(result[0]["embedding"])
    except Exception as e:
        print(f"  [WARN] Could not embed {img_path}: {e}")
    return None

# ── MATCHING ──────────────────────────────────────────────
def cosine_distance(a, b):
    a, b = np.array(a), np.array(b)
    denom = (np.linalg.norm(a) * np.linalg.norm(b))
    if denom == 0:
        return 1.0
    return 1 - np.dot(a, b) / denom

def identify_face(embedding, known_faces):
    best_name = "Unknown"
    best_dist = 1.0
    for name, embeddings in known_faces.items():
        for emb in embeddings:
            dist = cosine_distance(embedding, emb)
            if dist < best_dist:
                best_dist = dist
                best_name = name if dist < DISTANCE_THRESHOLD else "Unknown"
    return best_name, best_dist

# ── ON-SCREEN OVERLAY ─────────────────────────────────────
def draw_overlay(frame, detections, entry_log, exit_log, fps):
    H, W = frame.shape[:2]

    # Top bar
    cv2.rectangle(frame, (0,0), (W,55), (18,18,18), -1)
    cv2.putText(frame, "No-Stop Attendance System",
                (10,20), cv2.FONT_HERSHEY_SIMPLEX, 0.55, (0,220,120), 1)
    cv2.putText(frame, f"FPS: {fps:.1f}  |  Present: {len(entry_log)}",
                (10,44), cv2.FONT_HERSHEY_SIMPLEX, 0.45, (160,160,160), 1)
    cv2.putText(frame, datetime.now().strftime("%Y-%m-%d  %H:%M:%S"),
                (W-210,20), cv2.FONT_HERSHEY_SIMPLEX, 0.45, (160,160,160), 1)

    # Face boxes
    for (x, y, w, h, name, dist) in detections:
        color = (0,200,80) if name != "Unknown" else (0,60,220)
        cv2.rectangle(frame, (x,y), (x+w,y+h), color, 2)
        conf  = f"{(1-dist)*100:.0f}%" if name != "Unknown" else ""
        label = f"{name} {conf}" if name != "Unknown" else "Unknown"
        cv2.rectangle(frame, (x, y-26), (x+w, y), color, -1)
        cv2.putText(frame, label, (x+4, y-7),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.52, (0,0,0), 2)

    # Right panel — entry/exit log
    panel_x = W - 170
    cv2.rectangle(frame, (panel_x, 55), (W, 55 + len(entry_log)*46 + 10),
                  (22,22,22), -1)
    for i, name in enumerate(sorted(entry_log.keys())):
        y_base = 75 + i * 46
        # Entry line
        cv2.putText(frame, f"IN  {entry_log[name]}",
                    (panel_x+6, y_base),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.38, (0,200,80), 1)
        # Exit line
        exit_time = exit_log.get(name, "---")
        color_exit = (0,180,255) if exit_time != "---" else (100,100,100)
        cv2.putText(frame, f"OUT {exit_time}",
                    (panel_x+6, y_base+18),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.38, color_exit, 1)
        cv2.putText(frame, name[:18],
                    (panel_x+6, y_base+36),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.36, (200,200,200), 1)

# ── MAIN ──────────────────────────────────────────────────
def main():
    print("\n" + "="*55)
    print("  ATTENDANCE ENGINE — Entry + Exit Tracking")
    print("="*55)

    init_excel(EXCEL_FILE)

    print(f"\n  Loading known faces from '{KNOWN_FACES_DIR}/'...")
    known_faces = load_known_faces(KNOWN_FACES_DIR)

    if not known_faces:
        print("\n  [ERROR] No faces loaded! Check your known_faces/ folder.")
        print("  Expected structure:")
        print("    known_faces/")
        print("      Aditi_Sharma/")
        print("        front_01.jpg")
        print("        left_01.jpg")
        print("  Run enroll_face.py first.\n")
        return

    print(f"\n  Loaded: {list(known_faces.keys())}")
    print(f"  Opening camera {CAMERA_INDEX}...\n")

    cap = cv2.VideoCapture(CAMERA_INDEX)
    if not cap.isOpened():
        print(f"  [ERROR] Cannot open camera {CAMERA_INDEX}"); return

    cap.set(cv2.CAP_PROP_FRAME_WIDTH,  640)
    cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
    print("  Camera open! Press Q to quit.\n")

    face_cascade = cv2.CascadeClassifier(
        cv2.data.haarcascades + "haarcascade_frontalface_default.xml"
    )

    # Tracking state
    entry_log       = {}   # name → "HH:MM:SS" entry time string
    exit_log        = {}   # name → "HH:MM:SS" exit time string
    last_seen       = {}   # name → timestamp (when last detected)
    currently_inside = set()  # names currently visible / recently visible

    frame_count  = 0
    fps_timer    = time.time()
    fps          = 0.0
    last_detections = []

    while True:
        ret, frame = cap.read()
        if not ret:
            time.sleep(0.05); continue

        frame       = cv2.flip(frame, 1)
        frame_count += 1

        # FPS
        elapsed = time.time() - fps_timer
        if elapsed >= 1.0:
            fps         = frame_count / elapsed
            frame_count = 0
            fps_timer   = time.time()

        # ── Process frame ──────────────────────────────
        if frame_count % FRAME_SKIP == 0:
            gray  = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            faces = face_cascade.detectMultiScale(
                gray, scaleFactor=1.1, minNeighbors=5, minSize=(60,60)
            )

            last_detections  = []
            detected_names   = set()

            for (x, y, w, h) in faces:
                pad = 20
                x1,y1 = max(0,x-pad), max(0,y-pad)
                x2,y2 = min(frame.shape[1],x+w+pad), min(frame.shape[0],y+h+pad)
                crop  = frame[y1:y2, x1:x2]

                try:
                    result = DeepFace.represent(
                        img_path          = crop,
                        model_name        = MODEL_NAME,
                        detector_backend  = "skip",
                        enforce_detection = False
                    )
                    if not result:
                        last_detections.append((x,y,w,h,"Unknown",1.0))
                        continue

                    emb          = np.array(result[0]["embedding"])
                    name, dist   = identify_face(emb, known_faces)
                    last_detections.append((x,y,w,h,name,dist))

                    if name != "Unknown":
                        detected_names.add(name)
                        last_seen[name] = time.time()

                        # ── LOG ENTRY ──────────────────
                        if name not in entry_log:
                            now_str = datetime.now().strftime("%H:%M:%S")
                            entry_log[name] = now_str
                            currently_inside.add(name)
                            log_entry(EXCEL_FILE, name)

                        # Person seen again after exit — new entry
                        elif name in exit_log:
                            # Re-entered after leaving
                            del exit_log[name]
                            currently_inside.add(name)
                            print(f"  [RE-ENTRY] {name} returned")

                except Exception:
                    last_detections.append((x,y,w,h,"Unknown",1.0))

            # ── EXIT DETECTION ─────────────────────────
            now = time.time()
            for name in list(currently_inside):
                if name not in detected_names:
                    gone_for = now - last_seen.get(name, now)
                    if gone_for >= EXIT_TIMEOUT_SECONDS:
                        # Person has left
                        currently_inside.discard(name)
                        exit_time = datetime.now().strftime("%H:%M:%S")
                        exit_log[name] = exit_time
                        log_exit(EXCEL_FILE, name)

        # ── Draw UI ────────────────────────────────────
        draw_overlay(frame, last_detections, entry_log, exit_log, fps)

        # Show exit countdown for people about to be marked as exited
        now = time.time()
        y_offset = 60
        for name in currently_inside:
            if name not in {d[4] for d in last_detections}:
                gone = now - last_seen.get(name, now)
                remaining = max(0, EXIT_TIMEOUT_SECONDS - gone)
                cv2.putText(frame,
                            f"{name} leaving... ({remaining:.0f}s)",
                            (10, y_offset),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.45, (0,180,255), 1)
                y_offset += 20

        cv2.imshow("Attendance System — Entry & Exit", frame)

        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    # ── On quit — mark everyone still inside as exited ──
    print("\n  Closing... marking remaining people as exited.")
    for name in currently_inside:
        if name not in exit_log:
            log_exit(EXCEL_FILE, name)

    cap.release()
    cv2.destroyAllWindows()

    print(f"\n  Session Summary:")
    print(f"  Entries : {list(entry_log.keys())}")
    print(f"  Exits   : {list(exit_log.keys())}")
    print(f"  Excel   : {os.path.abspath(EXCEL_FILE)}\n")

if __name__ == "__main__":
    main()
